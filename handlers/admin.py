from telebot import types, TeleBot
from telebot.types import Message, CallbackQuery
from config import ADMIN_ID
from data.db import (
    get_pending_applications,
    clear_applications,
    update_application_lesson,
    get_application_by_id,
    format_date_for_display,
    validate_date_format,
    get_all_applications,
    get_all_archive,
    get_open_contacts,
    reply_to_contact,
    get_contact_by_id,
    ban_user_by_contact,
    get_all_reviews,
    get_review_stats,
    clear_reviews,
    get_all_courses,
    clear_courses,
    get_all_contacts,
    get_database_stats,
    migrate_database,
    clear_archive
)
from state.users import writing_ids
import utils.menu as menu
from openpyxl.utils import get_column_letter
import tempfile
import os
import re
import openpyxl
from utils.logger import setup_logger, log_bot_startup, log_bot_shutdown, log_error, log_admin_action
from utils.security import log_security_event
from utils.menu import create_admin_menu, create_confirm_menu

logger = setup_logger('admin')

def is_admin(user_id):
    return str(user_id) == str(ADMIN_ID)

def notify_admin_new_application(bot, application_data):
    try:
        notification = (
            "🔔 Новая заявка на обучение!\n\n"
            f"👤 Имя: {application_data.get('parent_name', 'Не указано')}\n"
            f"🧒 Ученик: {application_data.get('student_name', 'Не указано')}\n"
            f"📱 Контакт: {application_data.get('contact', 'Не указан')}\n"
            f"📚 Курс: {application_data.get('course', 'Не указан')}\n"
            f"📅 Возраст: {application_data.get('age', 'Не указан')}"
        )
        bot.send_message(ADMIN_ID, notification)
        print(f"[✅] Уведомление отправлено админу {ADMIN_ID}")
    except Exception as e:
        print(f"[❌] Ошибка при отправке уведомления админу: {str(e)}")


def register(bot, logger):
    @bot.message_handler(commands=["security_report"])
    def handle_security_report(message):
        if not is_admin(message.from_user.id):
            logger.warning(f"User {message.from_user.id} tried to access admin command security_report")
            return
        
        try:
            # Получаем отчет за последние 24 часа
            report = security_logger.get_security_report(hours=24)
            
            report_text = (
                "🔒 Отчет по безопасности (за последние 24 часа):\n\n"
                f"🚫 Неудачные попытки входа: {report.get('failed_logins', 0)}\n"
                f"⚠️ Подозрительная активность: {report.get('suspicious_activities', 0)}\n"
                f"⏱️ Превышения rate limit: {report.get('rate_limit_exceeded', 0)}\n"
                f"🚫 Заблокированные пользователи: {report.get('user_bans', 0)}\n"
                f"🚪 Несанкционированный доступ: {report.get('unauthorized_access', 0)}\n"
                f"❌ Ошибки валидации: {report.get('input_validation_failed', 0)}\n\n"
                "📊 Общая статистика безопасности:"
            )
            
            # Добавляем общую оценку безопасности
            total_events = sum(report.values())
            if total_events == 0:
                report_text += "\n✅ Все спокойно, угроз не обнаружено"
            elif total_events < 10:
                report_text += "\n🟡 Низкий уровень угроз"
            elif total_events < 50:
                report_text += "\n🟠 Средний уровень угроз"
            else:
                report_text += "\n🔴 Высокий уровень угроз!"
            
            bot.send_message(message.chat.id, report_text)
            logger.info(f"Admin {message.from_user.id} requested security report")
            
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при получении отчета безопасности: {str(e)}")
            logger.error(f"Error in security report: {e}")

    @bot.message_handler(commands=["ClearApplications"])
    def handle_clear_command(message):
        if not is_admin(message.from_user.id):
            logger.warning(f"User {message.from_user.id} tried to access admin command ClearApplications")
            return
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Да, очистить", callback_data="confirm_clear"),
            types.InlineKeyboardButton("❌ Нет", callback_data="cancel_clear")
        )
        bot.send_message(message.chat.id, "⚠️ Вы уверены, что хотите удалить все заявки?\nЭто действие необратимо.", reply_markup=markup)
        logger.info(f"Admin {message.from_user.id} initiated ClearApplications")

    @bot.message_handler(commands=["ClearArchive"])
    def handle_clear_archive_command(message):
        if not is_admin(message.from_user.id):
            logger.warning(f"User {message.from_user.id} tried to access admin command ClearArchive")
            return

        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Да, очистить архив", callback_data="confirm_clear_archive"),
            types.InlineKeyboardButton("❌ Нет", callback_data="cancel_clear_archive")
        )
        bot.send_message(
            message.chat.id,
            "⚠️ Вы уверены, что хотите удалить все архивные заявки?\nЭто действие необратимо.",
            reply_markup=markup
        )
        logger.info(f"Admin {message.from_user.id} initiated ClearArchive")

    @bot.callback_query_handler(func=lambda c: c.data in ["confirm_clear_archive", "cancel_clear_archive"])
    def handle_clear_archive_confirm(call):
        if not is_admin(call.from_user.id):
            logger.warning(f"User {call.from_user.id} tried to confirm archive clear without admin rights")
            return

        if call.data == "confirm_clear_archive":
            clear_archive()
            bot.send_message(call.message.chat.id, "🧹 Архив успешно очищен.")
            logger.info(f"Admin {call.from_user.id} cleared archive")
        else:
            bot.send_message(call.message.chat.id, "❌ Очистка архива отменена.")
            logger.info(f"Admin {call.from_user.id} cancelled archive clear")

    @bot.callback_query_handler(func=lambda call: call.data in ["confirm_clear", "cancel_clear"])
    def handle_clear_confirm(call):
        chat_id = call.message.chat.id
        if not is_admin(call.from_user.id):
            logger.warning(f"User {call.from_user.id} tried to confirm clear without admin rights")
            return

        if call.data == "confirm_clear":
            clear_applications()
            bot.send_message(chat_id, "✅ Все заявки успешно удалены.")
            logger.info(f"Admin {call.from_user.id} cleared all applications")
        else:
            bot.send_message(chat_id, "❌ Очистка отменена.")
            logger.info(f"Admin {call.from_user.id} cancelled clear")

    @bot.message_handler(func=lambda m: m.text == "📋 Список заявок" and is_admin(m.from_user.id))
    def handle_pending_applications(message):
        try:
            applications = get_pending_applications()
            if not applications:
                bot.send_message(message.chat.id, "✅ Нет заявок без назначенной даты")
                return

            for app in applications:
                app_id, tg_id, parent_name, student_name, age, contact, course, lesson_date, lesson_link, status, created_at, reminder_sent = app
                formatted_created = format_date_for_display(created_at)
                # Статус заявки
                if status == "Назначено":
                    status_str = "Назначено"
                else:
                    status_str = "Ожидает"
                text = (
                    f"🆔 Заявка #{app_id}\n"
                    f"👤 Родитель: {parent_name}\n"
                    f"🧒 Ученик: {student_name}\n"
                    f"📞 Контакт: {contact or 'не указан'}\n"
                    f"🎂 Возраст: {age}\n"
                    f"📘 Курс: {course}\n"
                    f"Статус: {status_str}\n"
                    f"🕒 Создано: {formatted_created}"
                )
                markup = types.InlineKeyboardMarkup()
                markup.add(types.InlineKeyboardButton("🕒 Назначить", callback_data=f"assign:{app_id}"))
                bot.send_message(message.chat.id, text, reply_markup=markup)
            logger.info(f"Admin {message.from_user.id} viewed pending applications")
        except Exception as e:
            bot.send_message(message.chat.id, f"❌ Ошибка при получении заявок: {str(e)}")
            logger.error(f"Error in handle_pending_applications: {e}")

    @bot.callback_query_handler(func=lambda c: c.data.startswith("assign:"))
    def handle_assign_callback(call):
        app_id = int(call.data.split(":")[1])
        writing_ids.add(call.from_user.id)
        bot.send_message(call.message.chat.id, f"📅 Введите дату и время урока для заявки #{app_id} (например: 22.06 17:00):", reply_markup=menu.get_cancel_button())
        bot.register_next_step_handler(call.message, lambda m: get_link(m, app_id))

    def get_link(message, app_id):
        if message.from_user.id not in writing_ids:
            return
        
        # Проверяем отмену
        if message.text == "🔙 Отмена":
            writing_ids.discard(message.from_user.id)
            menu.handle_cancel_action(bot, message, "урок", logger)
            return
        
        date_text = message.text.strip()
        
        # Валидируем дату
        is_valid, result = validate_date_format(date_text)
        
        if not is_valid:
            bot.send_message(
                message.chat.id, 
                f"❌ {result}\n\n📅 Попробуйте еще раз в формате ДД.ММ ЧЧ:ММ (например: 22.06 17:00):",
                reply_markup=menu.get_cancel_button()
            )
            bot.register_next_step_handler(message, lambda m: get_link(m, app_id))
            return
        
        # Сохраняем валидную дату
        user_data = getattr(message, '_user_data', {})
        user_data['valid_date'] = result
        message._user_data = user_data
        
        bot.send_message(message.chat.id, "🔗 Введите ссылку на урок:", reply_markup=menu.get_cancel_button())
        bot.register_next_step_handler(message, lambda m: finalize_lesson(m, app_id, date_text))

    def finalize_lesson(message, app_id, date_text):
        if message.from_user.id not in writing_ids:
            return
        
        # Проверяем отмену
        if message.text == "🔙 Отмена":
            writing_ids.discard(message.from_user.id)
            menu.handle_cancel_action(bot, message, "урок", logger)
            return
        
        link = message.text.strip()
        
        # Получаем валидированную дату
        user_data = getattr(message, '_user_data', {})
        valid_date = user_data.get('valid_date')
        
        if valid_date:
            # Используем валидированную дату (datetime объект)
            update_application_lesson(app_id, valid_date, link)
            formatted_date = format_date_for_display(valid_date)
        else:
            # Fallback - используем строку и надеемся на лучшее
            update_application_lesson(app_id, date_text, link)
            formatted_date = date_text
        
        bot.send_message(message.chat.id, f"✅ Урок назначен!\n📅 {formatted_date}\n🔗 {link}", reply_markup=menu.get_admin_menu())

        app = get_application_by_id(app_id)
        if app:
            tg_id = app[1]
            course = app[6]
            try:
                bot.send_message(
                    int(tg_id),
                    f"📅 Вам назначен урок!\n📘 Курс: {course}\n🗓 Дата: {formatted_date}\n🔗 Ссылка: {link}"
                )
                logger.info(f"Lesson notification sent to user {tg_id}")
            except Exception as e:
                logger.error(f"Failed to notify user {tg_id}: {e}")

        writing_ids.discard(message.from_user.id)
        logger.info(f"Admin {message.from_user.id} assigned lesson for application {app_id}")

    @bot.message_handler(func=lambda m: m.text == "📚 Редактировать курсы" and is_admin(m.from_user.id))
    def handle_course_menu(message):
        markup = menu.get_course_editor_menu()
        bot.send_message(message.chat.id, "🎓 Меню редактирования курсов:", reply_markup=markup)

    @bot.message_handler(func=lambda m: m.text == "⬇️ Выгрузить данные" and is_admin(m.from_user.id))
    def handle_export_data(message):
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("Заявки", callback_data="export_applications"),
            types.InlineKeyboardButton("Архив", callback_data="export_archive")
        )
        bot.send_message(message.chat.id, "Что выгрузить?", reply_markup=markup)

    @bot.callback_query_handler(func=lambda c: c.data in ["export_applications", "export_archive"])
    def handle_export_choice(call):
        if call.data == "export_applications":
            data = get_all_applications()
            filename = "applications_export.xlsx"
            headers = [
                "ID", "TG ID", "Родитель", "Ученик", "Возраст", "Контакт", "Курс",
                "Дата урока", "Ссылка", "Статус", "Создано"
            ]
            # Формируем строки с вычисленным статусом
            rows = []
            for row in data:
                app_id, tg_id, parent_name, student_name, age, contact, course, lesson_date, lesson_link, status, created_at, reminder_sent = row
                if status == "Назначено":
                    status_str = "Назначено"
                else:
                    status_str = "Ожидает"
                rows.append([
                    app_id, tg_id, parent_name, student_name, age, contact, course,
                    lesson_date, lesson_link, status_str, created_at
                ])
        else:
            data = get_all_archive()
            filename = "archive_export.xlsx"
            headers = [
                "ID", "TG ID", "Родитель", "Ученик", "Возраст", "Контакт", "Курс",
                "Дата урока", "Ссылка", "Статус", "Создано", "Кем отменено", "Комментарий"
            ]
            rows = data
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        # Автоширина столбцов
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2
        # Сохраняем во временный файл
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            bot.send_document(call.message.chat.id, f, visible_file_name=filename)
        os.remove(tmp_path)
        bot.answer_callback_query(call.id, "Выгрузка завершена")

    @bot.message_handler(func=lambda m: m.text == "📨 Обращения пользователей" and is_admin(m.from_user.id))
    def handle_contacts_menu(message):
        contacts = get_open_contacts()
        if not contacts:
            bot.send_message(message.chat.id, "✅ Нет новых обращений.", reply_markup=menu.get_admin_menu())
            return
        for c in contacts:
            contact_id, user_tg_id, user_contact, msg, admin_reply, status, created_at, reply_at, banned, ban_reason = c
            # Проверяем, есть ли вложение
            file_match = re.match(r"\[Вложение: (\w+), file_id: ([\w\-_]+)\](.*)", msg, re.DOTALL)
            if file_match:
                file_type, file_id, caption = file_match.groups()
                caption = caption.strip() or None
                if file_type == 'photo':
                    bot.send_photo(message.chat.id, file_id, caption=f"Обращение #{contact_id} от {user_contact}\n{caption or ''}")
                elif file_type == 'document':
                    bot.send_document(message.chat.id, file_id, caption=f"Обращение #{contact_id} от {user_contact}\n{caption or ''}")
                elif file_type == 'voice':
                    bot.send_voice(message.chat.id, file_id)
                elif file_type == 'video':
                    bot.send_video(message.chat.id, file_id, caption=f"Обращение #{contact_id} от {user_contact}\n{caption or ''}")
                elif file_type == 'video_note':
                    bot.send_video_note(message.chat.id, file_id)
                # Текстовое описание
                text = (
                    f"🆘 Обращение #{contact_id}\n"
                    f"👤 Пользователь: {user_contact} (ID: {user_tg_id})\n"
                    f"⏰ Время: {created_at}\n"
                    f"Статус: {status}\n"
                    f"\nТекст: (см. вложение выше)"
                )
            else:
                text = (
                    f"🆘 Обращение #{contact_id}\n"
                    f"👤 Пользователь: {user_contact} (ID: {user_tg_id})\n"
                    f"⏰ Время: {created_at}\n"
                    f"Статус: {status}\n"
                    f"\nТекст: {msg}"
                )
            markup = types.InlineKeyboardMarkup()
            markup.add(types.InlineKeyboardButton("✉️ Ответить", callback_data=f"reply_contact:{contact_id}"))
            markup.add(types.InlineKeyboardButton("🚫 Забанить", callback_data=f"ban_contact:{user_tg_id}"))
            bot.send_message(message.chat.id, text, reply_markup=markup)

    @bot.callback_query_handler(func=lambda c: c.data.startswith("reply_contact:"))
    def handle_reply_contact(call):
        contact_id = int(call.data.split(":")[1])
        bot.send_message(call.message.chat.id, "✍️ Введите ответ пользователю:", reply_markup=menu.get_cancel_button())
        bot.register_next_step_handler(call.message, lambda m: process_admin_reply(m, contact_id))

    def process_admin_reply(message, contact_id):
        if message.text == "🔙 Отмена":
            bot.send_message(message.chat.id, "Ответ отменён.", reply_markup=menu.get_admin_menu())
            return
        reply_to_contact(contact_id, message.text)
        contact = get_contact_by_id(contact_id)
        user_tg_id = contact[1]
        bot.send_message(message.chat.id, "✅ Ответ отправлен пользователю.", reply_markup=menu.get_admin_menu())
        bot.send_message(int(user_tg_id), f"📨 Ответ от администратора:\n\n{message.text}")

    @bot.callback_query_handler(func=lambda c: c.data.startswith("ban_contact:"))
    def handle_ban_contact(call):
        user_tg_id = call.data.split(":")[1]
        bot.send_message(call.message.chat.id, "✍️ Введите причину бана пользователя:", reply_markup=menu.get_cancel_button())
        bot.register_next_step_handler(call.message, lambda m: process_ban_reason(m, user_tg_id))

    def process_ban_reason(message, user_tg_id):
        if message.text == "🔙 Отмена":
            bot.send_message(message.chat.id, "Бан отменён.", reply_markup=menu.get_admin_menu())
            return
        reason = message.text.strip()
        from data.db import ban_user_by_contact
        ban_user_by_contact(user_tg_id, reason)
        bot.send_message(message.chat.id, f"Пользователь {user_tg_id} заблокирован для обращений.\nПричина: {reason}", reply_markup=menu.get_admin_menu())
        try:
            bot.send_message(int(user_tg_id), f"🚫 Вы были заблокированы для обращений к админу.\nПричина: {reason}")
        except Exception as e:
            logger.error(f"Не удалось уведомить пользователя о бане: {e}")

    @bot.message_handler(commands=["ClearContacts"])
    def handle_clear_contacts(message):
        if not is_admin(message.from_user.id):
            logger.warning(f"User {message.from_user.id} tried to access admin command ClearContacts")
            return
        from data.db import clear_contacts
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Да, очистить обращения", callback_data="confirm_clear_contacts"),
            types.InlineKeyboardButton("❌ Нет", callback_data="cancel_clear_contacts")
        )
        bot.send_message(message.chat.id, "⚠️ Вы уверены, что хотите удалить все обращения пользователей? Это действие необратимо.", reply_markup=markup)
        logger.info(f"Admin {message.from_user.id} initiated ClearContacts")

    @bot.callback_query_handler(func=lambda c: c.data in ["confirm_clear_contacts", "cancel_clear_contacts"])
    def handle_clear_contacts_confirm(call):
        if not is_admin(call.from_user.id):
            logger.warning(f"User {call.from_user.id} tried to confirm clear contacts without admin rights")
            return
        from data.db import clear_contacts
        if call.data == "confirm_clear_contacts":
            clear_contacts()
            bot.send_message(call.message.chat.id, "🧹 Все обращения пользователей успешно удалены.")
            logger.info(f"Admin {call.from_user.id} cleared all contacts")
        else:
            bot.send_message(call.message.chat.id, "❌ Очистка обращений отменена.")
            logger.info(f"Admin {call.from_user.id} cancelled contacts clear")

    # === ОБРАБОТЧИКИ ДЛЯ ОТЗЫВОВ ===
    
    @bot.message_handler(func=lambda m: m.text == "⭐ Отзывы" and is_admin(m.from_user.id))
    def handle_reviews_menu(message):
        """Показывает меню управления отзывами"""
        try:
            stats = get_review_stats()
            total_reviews, avg_rating, positive_reviews, negative_reviews = stats
            
            # Проверяем, что avg_rating не None
            avg_rating_display = f"{avg_rating:.1f}" if avg_rating is not None else "0.0"
            
            msg = (
                "⭐ Управление отзывами\n\n"
                f"📊 Статистика:\n"
                f"• Всего отзывов: {total_reviews}\n"
                f"• Средняя оценка: {avg_rating_display}/10\n"
                f"• Положительных (8-10): {positive_reviews}\n"
                f"• Отрицательных (1-5): {negative_reviews}\n\n"
                "Выберите действие:\n\n"
                "💡 Для очистки всех отзывов используйте команду /ClearReviews"
            )
            
            markup = types.InlineKeyboardMarkup()
            markup.add(
                types.InlineKeyboardButton("📋 Все отзывы", callback_data="view_all_reviews"),
                types.InlineKeyboardButton("📊 Статистика", callback_data="review_stats")
            )
            
            bot.send_message(message.chat.id, msg, reply_markup=markup)
            logger.info(f"Admin {message.from_user.id} opened reviews menu")
            
        except Exception as e:
            logger.error(f"Error in handle_reviews_menu: {e}")
    
    @bot.callback_query_handler(func=lambda c: c.data == "view_all_reviews")
    def handle_view_all_reviews(call):
        """Показывает все отзывы"""
        try:
            reviews = get_all_reviews()
            if not reviews:
                bot.send_message(call.message.chat.id, "📭 Пока нет отзывов.")
                return
            batch_size = 10
            for batch_start in range(0, len(reviews), batch_size):
                msg = "📋 Все отзывы:\n\n"
                for i, review in enumerate(reviews[batch_start:batch_start+batch_size], batch_start+1):
                    review_id, rating, feedback, is_anonymous, parent_name, student_name, course, created_at, user_tg_id = review
                    from datetime import datetime
                    try:
                        dt = datetime.fromisoformat(created_at)
                        date_str = dt.strftime("%d.%m.%Y %H:%M")
                    except:
                        date_str = "недавно"
                    stars = "⭐" * rating
                    
                    # Определяем автора отзыва
                    if is_anonymous:
                        author = "Анонимный"
                    elif parent_name is None and student_name is None and course is None:
                        # Заявка удалена, но отзыв не анонимный - показываем ID пользователя
                        author = f"[Заявка удалена] ID: {user_tg_id}"
                    else:
                        author = f"{parent_name} ({student_name})"
                    
                    # Определяем курс
                    if course is None:
                        course = "[Заявка удалена]"
                    
                    feedback_display = feedback[:80] + ("..." if feedback and len(feedback) > 80 else "")
                    if not feedback_display.strip():
                        feedback_display = "Без текста, только оценка"
                    
                    msg += (
                        f"{i}. {stars} ({rating}/10)\n"
                        f"📘 Курс: {course}\n"
                        f"👤 {author}\n"
                        f"📝 {feedback_display}\n"
                        f"📅 {date_str}\n"
                        f"🆔 ID: {review_id}\n\n"
                    )
                bot.send_message(call.message.chat.id, msg)
            logger.info(f"Admin {call.from_user.id} viewed all reviews")
        except Exception as e:
            logger.error(f"Error in handle_view_all_reviews: {e}")
    
    @bot.callback_query_handler(func=lambda c: c.data == "review_stats")
    def handle_review_stats(call):
        """Показывает детальную статистику отзывов"""
        try:
            stats = get_review_stats()
            total_reviews, avg_rating, positive_reviews, negative_reviews = stats
            
            if total_reviews == 0:
                bot.send_message(call.message.chat.id, "📭 Пока нет отзывов для анализа.")
                return
            
            # Вычисляем проценты
            positive_percent = (positive_reviews / total_reviews) * 100 if total_reviews > 0 else 0
            negative_percent = (negative_reviews / total_reviews) * 100 if total_reviews > 0 else 0
            neutral_percent = 100 - positive_percent - negative_percent
            
            # Проверяем, что avg_rating не None
            avg_rating_display = f"{avg_rating:.1f}" if avg_rating is not None else "0.0"
            avg_rating_value = avg_rating if avg_rating is not None else 0
            
            msg = (
                "📊 Детальная статистика отзывов\n\n"
                f"📈 Общие показатели:\n"
                f"• Всего отзывов: {total_reviews}\n"
                f"• Средняя оценка: {avg_rating_display}/10\n\n"
                f"📊 Распределение:\n"
                f"• Отлично (8-10): {positive_reviews} ({positive_percent:.1f}%)\n"
                f"• Хорошо (6-7): {total_reviews - positive_reviews - negative_reviews} ({neutral_percent:.1f}%)\n"
                f"• Плохо (1-5): {negative_reviews} ({negative_percent:.1f}%)\n\n"
                f"🎯 Рекомендации:\n"
            )
            
            if avg_rating_value >= 8:
                msg += "🌟 Отличные результаты! Продолжайте в том же духе!"
            elif avg_rating_value >= 6:
                msg += "👍 Хорошие результаты. Есть возможности для улучшения."
            else:
                msg += "⚠️ Требуется внимание к качеству услуг."
            
            bot.send_message(call.message.chat.id, msg)
            logger.info(f"Admin {call.from_user.id} viewed review statistics")
            
        except Exception as e:
            logger.error(f"Error in handle_review_stats: {e}")
    
    @bot.message_handler(commands=["ClearReviews"])
    def handle_clear_reviews_command(message):
        """Команда для очистки отзывов"""
        if not is_admin(message.from_user.id):
            logger.warning(f"User {message.from_user.id} tried to access admin command ClearReviews")
            return
        
        markup = types.InlineKeyboardMarkup()
        markup.add(
            types.InlineKeyboardButton("✅ Да, очистить", callback_data="confirm_clear_reviews"),
            types.InlineKeyboardButton("❌ Нет", callback_data="cancel_clear_reviews")
        )
        bot.send_message(
            message.chat.id,
            "⚠️ Вы уверены, что хотите удалить все отзывы?\nЭто действие необратимо.",
            reply_markup=markup
        )
        logger.info(f"Admin {message.from_user.id} initiated ClearReviews command")

    @bot.callback_query_handler(func=lambda c: c.data in ["confirm_clear_reviews", "cancel_clear_reviews"])
    def handle_clear_reviews_confirm(call):
        """Подтверждает или отменяет очистку отзывов"""
        try:
            if not is_admin(call.from_user.id):
                logger.warning(f"User {call.from_user.id} tried to confirm clear reviews without admin rights")
                return
                
            if call.data == "confirm_clear_reviews":
                clear_reviews()
                bot.send_message(call.message.chat.id, "🧹 Все отзывы успешно удалены.")
                logger.info(f"Admin {call.from_user.id} cleared all reviews")
            else:
                bot.send_message(call.message.chat.id, "❌ Очистка отзывов отменена.")
                logger.info(f"Admin {call.from_user.id} cancelled reviews clear")
                
        except Exception as e:
            logger.error(f"Error in handle_clear_reviews_confirm: {e}")

def register_admin_handlers(bot: TeleBot):
    @bot.message_handler(commands=['admin'])
    def admin_command(message: Message):
        if str(message.from_user.id) != ADMIN_ID:
            log_security_event(f"Unauthorized admin access attempt by {message.from_user.id}")
            return
        
        log_admin_action(logger, message.from_user.id, "Admin panel accessed")
        bot.reply_to(message, "🔐 Панель администратора", reply_markup=create_admin_menu())

    @bot.callback_query_handler(func=lambda call: call.data.startswith('admin_'))
    def handle_admin_callback(call: CallbackQuery):
        if str(call.from_user.id) != ADMIN_ID:
            log_security_event(f"Unauthorized admin callback by {call.from_user.id}")
            return
        
        action = call.data.split('_')[1]
        log_admin_action(logger, call.from_user.id, f"Admin action '{action}'")
        
        if action == 'applications':
            show_applications(bot, call.message)
        elif action == 'archive':
            show_archive(bot, call.message)
        elif action == 'courses':
            show_courses(bot, call.message)
        elif action == 'contacts':
            show_contacts(bot, call.message)
        elif action == 'reviews':
            show_reviews(bot, call.message)
        elif action == 'clear_applications':
            confirm_clear_applications(bot, call.message)
        elif action == 'clear_archive':
            confirm_clear_archive(bot, call.message)
        elif action == 'clear_courses':
            confirm_clear_courses(bot, call.message)
        elif action == 'clear_contacts':
            confirm_clear_contacts(bot, call.message)
        elif action == 'clear_reviews':
            confirm_clear_reviews(bot, call.message)
        elif action == 'db_stats':
            show_database_stats(bot, call.message)
        elif action == 'migrate_db':
            migrate_database_action(bot, call.message)
        
        bot.answer_callback_query(call.id)

def show_applications(bot: TeleBot, message: Message):
    applications = get_all_applications()
    if not applications:
        bot.edit_message_text("📝 Нет активных заявок", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "📝 **Активные заявки:**\n\n"
    for app in applications[:10]:  # Показываем только первые 10
        text += f"ID: {app[0]}\n"
        text += f"Родитель: {app[2]}\n"
        text += f"Ученик: {app[3]} ({app[4]} лет)\n"
        text += f"Курс: {app[6]}\n"
        text += f"Статус: {app[10]}\n"
        text += f"Дата создания: {app[11]}\n"
        text += "─" * 30 + "\n"
    
    if len(applications) > 10:
        text += f"\n... и еще {len(applications) - 10} заявок"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def show_archive(bot: TeleBot, message: Message):
    archive = get_all_archive()
    if not archive:
        bot.edit_message_text("🗄️ Архив пуст", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "🗄️ **Архив:**\n\n"
    for record in archive[:10]:  # Показываем только первые 10
        text += f"ID: {record[0]}\n"
        text += f"Родитель: {record[2]}\n"
        text += f"Ученик: {record[3]} ({record[4]} лет)\n"
        text += f"Курс: {record[6]}\n"
        text += f"Статус: {record[10]}\n"
        text += f"Дата архивирования: {record[12]}\n"
        text += "─" * 30 + "\n"
    
    if len(archive) > 10:
        text += f"\n... и еще {len(archive) - 10} записей"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def show_courses(bot: TeleBot, message: Message):
    courses = get_all_courses()
    if not courses:
        bot.edit_message_text("📚 Нет курсов", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "📚 **Курсы:**\n\n"
    for course in courses:
        status = "✅ Активен" if course[3] else "❌ Неактивен"
        text += f"ID: {course[0]}\n"
        text += f"Название: {course[1]}\n"
        text += f"Описание: {course[2]}\n"
        text += f"Статус: {status}\n"
        text += "─" * 30 + "\n"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def show_contacts(bot: TeleBot, message: Message):
    contacts = get_all_contacts()
    if not contacts:
        bot.edit_message_text("📞 Нет обращений", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "📞 **Обращения:**\n\n"
    for contact in contacts[:10]:  # Показываем только первые 10
        text += f"ID: {contact[0]}\n"
        text += f"Пользователь: {contact[1]}\n"
        text += f"Контакты: {contact[2]}\n"
        text += f"Сообщение: {contact[3][:50]}...\n"
        text += f"Статус: {contact[5]}\n"
        text += f"Дата: {contact[6]}\n"
        text += "─" * 30 + "\n"
    
    if len(contacts) > 10:
        text += f"\n... и еще {len(contacts) - 10} обращений"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def show_reviews(bot: TeleBot, message: Message):
    reviews = get_all_reviews()
    if not reviews:
        bot.edit_message_text("⭐ Нет отзывов", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "⭐ **Отзывы:**\n\n"
    for review in reviews[:10]:  # Показываем только первые 10
        text += f"ID: {review[0]}\n"
        text += f"Рейтинг: {review[1]}/10\n"
        text += f"Отзыв: {review[2][:50]}...\n"
        text += f"Анонимный: {'Да' if review[3] else 'Нет'}\n"
        text += f"Родитель: {review[4]}\n"
        text += f"Ученик: {review[5]}\n"
        text += f"Курс: {review[6]}\n"
        text += f"Дата: {review[7]}\n"
        text += "─" * 30 + "\n"
    
    if len(reviews) > 10:
        text += f"\n... и еще {len(reviews) - 10} отзывов"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def show_database_stats(bot: TeleBot, message: Message):
    """Показывает статистику базы данных"""
    stats = get_database_stats()
    if not stats:
        bot.edit_message_text("❌ Не удалось получить статистику БД", chat_id=message.chat.id, message_id=message.message_id)
        return
    
    text = "📊 **Статистика базы данных:**\n\n"
    text += f"📝 Заявок: {stats.get('applications_count', 0)}\n"
    text += f"📚 Курсов: {stats.get('courses_count', 0)}\n"
    text += f"📞 Обращений: {stats.get('contacts_count', 0)}\n"
    text += f"⭐ Отзывов: {stats.get('reviews_count', 0)}\n"
    text += f"🗄️ В архиве: {stats.get('archive_count', 0)}\n"
    text += f"💾 Размер БД: {stats.get('database_size_mb', 0)} МБ\n"
    text += f"🔍 Индексов в applications: {stats.get('applications_indexes', 0)}\n"
    
    bot.edit_message_text(text, chat_id=message.chat.id, message_id=message.message_id, parse_mode='Markdown')

def migrate_database_action(bot: TeleBot, message: Message):
    """Выполняет миграцию базы данных"""
    try:
        if migrate_database():
            bot.edit_message_text("✅ Миграция базы данных выполнена успешно!", chat_id=message.chat.id, message_id=message.message_id)
        else:
            bot.edit_message_text("❌ Ошибка при выполнении миграции БД", chat_id=message.chat.id, message_id=message.message_id)
    except Exception as e:
        logger.error(f"Error during database migration: {e}")
        bot.edit_message_text(f"❌ Ошибка при миграции: {str(e)}", chat_id=message.chat.id, message_id=message.message_id)

def confirm_clear_applications(bot: TeleBot, message: Message):
    """Показывает меню подтверждения очистки заявок"""
    bot.edit_message_text(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ активные заявки!\n"
        "Это действие нельзя отменить.\n\n"
        "Продолжить?",
        chat_id=message.chat.id,
        message_id=message.message_id,
        parse_mode='Markdown',
        reply_markup=create_confirm_menu('clear_applications')
    )

def confirm_clear_archive(bot: TeleBot, message: Message):
    """Показывает меню подтверждения очистки архива"""
    bot.edit_message_text(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ записи из архива!\n"
        "Это действие нельзя отменить.\n\n"
        "Продолжить?",
        chat_id=message.chat.id,
        message_id=message.message_id,
        parse_mode='Markdown',
        reply_markup=create_confirm_menu('clear_archive')
    )

def confirm_clear_courses(bot: TeleBot, message: Message):
    """Показывает меню подтверждения очистки курсов"""
    bot.edit_message_text(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ курсы!\n"
        "Это действие нельзя отменить.\n\n"
        "Продолжить?",
        chat_id=message.chat.id,
        message_id=message.message_id,
        parse_mode='Markdown',
        reply_markup=create_confirm_menu('clear_courses')
    )

def confirm_clear_contacts(bot: TeleBot, message: Message):
    """Показывает меню подтверждения очистки обращений"""
    bot.edit_message_text(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ обращения пользователей!\n"
        "Это действие нельзя отменить.\n\n"
        "Продолжить?",
        chat_id=message.chat.id,
        message_id=message.message_id,
        parse_mode='Markdown',
        reply_markup=create_confirm_menu('clear_contacts')
    )

def confirm_clear_reviews(bot: TeleBot, message: Message):
    """Показывает меню подтверждения очистки отзывов"""
    bot.edit_message_text(
        "⚠️ **ВНИМАНИЕ!**\n\n"
        "Вы собираетесь удалить ВСЕ отзывы!\n"
        "Это действие нельзя отменить.\n\n"
        "Продолжить?",
        chat_id=message.chat.id,
        message_id=message.message_id,
        parse_mode='Markdown',
        reply_markup=create_confirm_menu('clear_reviews')
    )

def register_handlers(bot):
    """Регистрация всех админских обработчиков"""
    logger = setup_logger('admin')
    register(bot, logger)  # Старые обработчики
    register_admin_handlers(bot)  # Новые инлайн-обработчики
